import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

def normalize_name(name: str) -> str:
    """
    Normalize student name using first and last name.
    """
    if pd.isna(name):
        return ""
    parts = str(name).strip().split()
    return parts[0].lower() if len(parts) == 1 else f"{parts[0]} {parts[-1]}".lower()


def reconcile_coscholastics(prev_file, master_file, output_file):
    """
    Reconcile previous activity allocations with next grade master list.
    """

    prev_grade = pd.read_excel(prev_file)
    next_grade = pd.read_excel(master_file)

    prev_grade["Normalized"] = prev_grade["Student Name"].apply(normalize_name)
    next_grade["Normalized"] = next_grade["Name of Student"].apply(normalize_name)

    prev_grade = prev_grade.rename(columns={
        "Student Name": "Name of Student",
        "Term I Activity": "Term I Activity",
        "Term II Activity": "Term II Activity"
    })

    merged = pd.merge(
        next_grade,
        prev_grade[["Normalized", "Term I Activity", "Term II Activity"]],
        on="Normalized",
        how="left"
    )

    merged["Form Submission"] = merged["Term I Activity"].apply(
        lambda x: "Submitted" if pd.notna(x) else "Not Submitted"
    )

    merged["Term I Activity"] = merged["Term I Activity"].fillna("None")
    merged["Term II Activity"] = merged["Term II Activity"].fillna("None")

    merged["Section"] = merged["Grade"].astype(str).str[-1]

    merged = merged[[
        "Sl No.", "Admission No.", "Name of Student", "Section",
        "Form Submission", "Term I Activity", "Term II Activity"
    ]]

    prev_names = set(prev_grade["Normalized"])
    next_names = set(next_grade["Normalized"])

    left_school = prev_grade[prev_grade["Normalized"].isin(prev_names - next_names)].copy()
    left_school["Sl No."] = None
    left_school["Admission No."] = "Left School"
    left_school["Section"] = "N/A"
    left_school["Form Submission"] = "Submitted"

    left_school = left_school[[
        "Sl No.", "Admission No.", "Name of Student", "Section",
        "Form Submission", "Term I Activity", "Term II Activity"
    ]]

    final_output = pd.concat([merged, left_school], ignore_index=True)
    final_output.to_excel(output_file, index=False)

    wb = load_workbook(output_file)
    ws = wb.active

    for row in range(2, ws.max_row + 1):
        if ws[f"E{row}"].value == "Not Submitted" or ws[f"B{row}"].value == "Left School":
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).font = Font(color="FF0000")

    wb.save(output_file)

    print(f"Reconciled file created → {output_file}")


if __name__ == "__main__":
    prev_file = input("Enter previous allocation file: ")
    master_file = input("Enter master file: ")
    output_file = input("Enter output file name: ")

    reconcile_coscholastics(prev_file, master_file, output_file)
